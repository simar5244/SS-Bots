#!/usr/bin/env python3
"""
VPAT to Scorecard Integration
Populates the UTA VPAT Scorecard Excel file with data from parsed VPATs
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
from typing import List, Dict
from vpat_parser import VPATParser, WCAGCriterion, VPATMetadata, ScorecardMapping
from datetime import datetime


class ScorecardPopulator:
    """Populate Excel scorecard with VPAT data"""
    
    # Mapping of WCAG criteria IDs to Excel row numbers (Sheet 3: ACR)
    # This would need to be customized based on the actual scorecard layout
    WCAG_ROW_MAPPING = {
        '1.1.1': 5,
        '1.2.1': 6,
        '1.2.2': 7,
        '1.2.3': 8,
        '1.2.4': 9,
        '1.2.5': 10,
        '1.3.1': 11,
        '1.3.2': 12,
        '1.3.3': 13,
        '1.3.4': 14,
        '1.3.5': 15,
        '1.4.1': 16,
        '1.4.2': 17,
        '1.4.3': 18,
        '1.4.4': 19,
        '1.4.5': 20,
        '1.4.10': 21,
        '1.4.11': 22,
        '1.4.12': 23,
        '1.4.13': 24,
        '2.1.1': 25,
        '2.1.2': 26,
        '2.1.4': 27,
        '2.2.1': 28,
        '2.2.2': 29,
        '2.3.1': 30,
        '2.4.1': 31,
        '2.4.2': 32,
        '2.4.3': 33,
        '2.4.4': 34,
        '2.4.5': 35,
        '2.4.6': 36,
        '2.4.7': 37,
        '2.5.1': 38,
        '2.5.2': 39,
        '2.5.3': 40,
        '2.5.4': 41,
        '3.1.1': 42,
        '3.1.2': 43,
        '3.2.1': 44,
        '3.2.2': 45,
        '3.2.3': 46,
        '3.2.4': 47,
        '3.3.1': 48,
        '3.3.2': 49,
        '3.3.3': 50,
        '3.3.4': 51,
        '4.1.1': 52,
        '4.1.2': 53,
        '4.1.3': 54,
    }
    
    # Column mappings for ACR sheet
    COLUMNS = {
        'criterion_id': 'A',
        'criterion_name': 'B',
        'conformance': 'C',  # Dropdown: Supports, Partially Supports, Does Not Support
        'impact': 'D',
        'notes': 'E',
    }
    
    def __init__(self, scorecard_path: str):
        """Initialize with path to scorecard template"""
        self.scorecard_path = scorecard_path
        self.workbook = None
        self.acr_sheet = None
        self.score_sheet = None
    
    def load_scorecard(self):
        """Load the Excel scorecard"""
        try:
            self.workbook = openpyxl.load_workbook(self.scorecard_path)
            
            # Get the ACR sheet (Sheet 3)
            if 'ACR' in self.workbook.sheetnames:
                self.acr_sheet = self.workbook['ACR']
            else:
                # Try to get by index (Sheet 3 = index 2)
                self.acr_sheet = self.workbook.worksheets[2]
            
            # Get the Score sheet (Sheet 2)
            if 'Score' in self.workbook.sheetnames:
                self.score_sheet = self.workbook['Score']
            else:
                self.score_sheet = self.workbook.worksheets[1]
            
            return True
        except Exception as e:
            raise Exception(f"Error loading scorecard: {str(e)}")
    
    def populate_metadata(self, metadata: VPATMetadata):
        """Populate metadata in the Score sheet"""
        if not self.score_sheet:
            return
        
        # Table 1: Resource details (adjust cell references as needed)
        # These would need to be customized based on actual scorecard layout
        try:
            # Example placements - adjust based on actual scorecard
            if metadata.product_name:
                self.score_sheet['B2'] = metadata.product_name
            
            if metadata.report_date:
                self.score_sheet['B3'] = metadata.report_date
            
            if metadata.wcag_version:
                self.score_sheet['B4'] = f"WCAG {metadata.wcag_version} Level {metadata.wcag_level or 'AA'}"
        except Exception as e:
            print(f"Warning: Could not populate metadata: {str(e)}")
    
    def populate_criteria(self, criteria: List[WCAGCriterion]):
        """Populate WCAG criteria in the ACR sheet"""
        if not self.acr_sheet:
            return
        
        populated_count = 0
        skipped_count = 0
        
        for criterion in criteria:
            # Get the row number for this criterion
            row = self.WCAG_ROW_MAPPING.get(criterion.criterion_id)
            
            if not row:
                skipped_count += 1
                continue
            
            # Populate conformance level (Column C)
            conformance_value = self._get_dropdown_value(criterion.scorecard_equivalent)
            self.acr_sheet[f'C{row}'] = conformance_value
            
            # Optionally add remarks to notes column (Column E)
            if criterion.remarks:
                current_notes = self.acr_sheet[f'E{row}'].value or ""
                vpat_note = f"VPAT: {criterion.conformance_level.value}"
                if criterion.remarks:
                    vpat_note += f" - {criterion.remarks[:100]}"  # Truncate long remarks
                
                if current_notes:
                    self.acr_sheet[f'E{row}'] = f"{current_notes}\n{vpat_note}"
                else:
                    self.acr_sheet[f'E{row}'] = vpat_note
            
            # Highlight cells that were auto-populated
            self._highlight_cell(f'C{row}', 'FFFFE0')  # Light yellow
            
            populated_count += 1
        
        return populated_count, skipped_count
    
    def _get_dropdown_value(self, scorecard_mapping: ScorecardMapping) -> str:
        """Convert scorecard mapping to dropdown value"""
        # The dropdown accepts: s, p, d (or full text)
        mapping = {
            ScorecardMapping.SUPPORTS: 's',
            ScorecardMapping.PARTIALLY_SUPPORTS: 'p',
            ScorecardMapping.DOES_NOT_SUPPORT: 'd',
        }
        return mapping.get(scorecard_mapping, 's')
    
    def _highlight_cell(self, cell_ref: str, color: str):
        """Highlight a cell with the given color"""
        try:
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            self.acr_sheet[cell_ref].fill = fill
        except Exception:
            pass  # Ignore styling errors
    
    def add_processing_note(self, validation_status: str):
        """Add a note about automated processing"""
        if not self.acr_sheet:
            return
        
        # Add note at the top of the sheet (adjust as needed)
        note = f"Auto-populated from VPAT on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        note += f"\nValidation Status: {validation_status}"
        
        try:
            # Add to a notes area (adjust cell reference as needed)
            self.acr_sheet['A1'] = note
            self.acr_sheet['A1'].font = Font(italic=True, size=9)
        except Exception:
            pass
    
    def save(self, output_path: str = None):
        """Save the populated scorecard"""
        if not self.workbook:
            raise Exception("No workbook loaded")
        
        save_path = output_path or self.scorecard_path
        self.workbook.save(save_path)
        return save_path
    
    def close(self):
        """Close the workbook"""
        if self.workbook:
            self.workbook.close()


def process_vpat_to_scorecard(vpat_pdf_path: str, scorecard_template_path: str, output_path: str = None):
    """
    Complete workflow: Parse VPAT, validate, and populate scorecard
    
    Args:
        vpat_pdf_path: Path to VPAT PDF file
        scorecard_template_path: Path to Excel scorecard template
        output_path: Optional output path for populated scorecard
    
    Returns:
        Dict with processing results
    """
    results = {
        'success': False,
        'validation_passed': False,
        'criteria_populated': 0,
        'criteria_skipped': 0,
        'errors': [],
        'warnings': [],
        'output_file': None,
    }
    
    try:
        # Step 1: Parse VPAT
        print("Step 1: Parsing VPAT PDF...")
        parser = VPATParser()
        metadata, criteria = parser.parse_pdf(vpat_pdf_path)
        print(f"  ✓ Extracted {len(criteria)} WCAG criteria")
        
        # Step 2: Validate
        print("\nStep 2: Validating VPAT...")
        validation = parser.validate(metadata)
        results['validation_passed'] = validation.is_valid
        results['errors'] = validation.errors
        results['warnings'] = validation.warnings
        
        if validation.is_valid:
            print("  ✓ VPAT validation passed")
        else:
            print("  ⚠️  VPAT validation failed - proceeding with manual review flag")
            for error in validation.errors:
                print(f"    - {error}")
        
        # Step 3: Populate scorecard
        print("\nStep 3: Populating scorecard...")
        populator = ScorecardPopulator(scorecard_template_path)
        populator.load_scorecard()
        
        # Populate metadata
        populator.populate_metadata(metadata)
        print("  ✓ Metadata populated")
        
        # Populate criteria
        populated, skipped = populator.populate_criteria(criteria)
        results['criteria_populated'] = populated
        results['criteria_skipped'] = skipped
        print(f"  ✓ Populated {populated} criteria ({skipped} skipped)")
        
        # Add processing note
        status = "VALID - Auto-processed" if validation.is_valid else "INVALID - Requires manual review"
        populator.add_processing_note(status)
        
        # Step 4: Save
        if not output_path:
            # Generate output filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            product_name = metadata.product_name or "Unknown"
            safe_name = "".join(c for c in product_name if c.isalnum() or c in (' ', '-', '_'))[:50]
            output_path = f"Scorecard_{safe_name}_{timestamp}.xlsx"
        
        output_file = populator.save(output_path)
        populator.close()
        
        results['output_file'] = output_file
        results['success'] = True
        
        print(f"\n✓ Scorecard saved to: {output_file}")
        
        # Step 5: Generate report
        print("\n" + "=" * 80)
        report = parser.generate_report(validation, criteria)
        print(report)
        
        return results
        
    except Exception as e:
        results['errors'].append(str(e))
        print(f"\n❌ Error: {str(e)}")
        return results


def main():
    """Command-line interface"""
    import sys
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Parse VPAT PDF and populate UTA VPAT Scorecard'
    )
    parser.add_argument('vpat_pdf', help='Path to VPAT PDF file')
    parser.add_argument('scorecard_template', help='Path to Excel scorecard template')
    parser.add_argument('-o', '--output', help='Output path for populated scorecard')
    
    args = parser.parse_args()
    
    results = process_vpat_to_scorecard(
        args.vpat_pdf,
        args.scorecard_template,
        args.output
    )
    
    if results['success']:
        print("\n✅ Processing completed successfully!")
        if not results['validation_passed']:
            print("⚠️  Note: VPAT validation failed - manual review required")
        sys.exit(0)
    else:
        print("\n❌ Processing failed!")
        for error in results['errors']:
            print(f"  - {error}")
        sys.exit(1)


if __name__ == "__main__":
    main()
